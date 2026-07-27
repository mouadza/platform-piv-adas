import openpyxl


def normalize_text(value):
    if value is None:
        return ""

    return str(value).strip().lower().replace("’", "'")


def extract_nom_gamme_from_excel(file):
    """
    Cherche la cellule contenant 'Gamme d'Essai'
    puis retourne la valeur située à droite ou en dessous.
    """

    try:
        if not file.name.lower().endswith((".xlsx", ".xlsm")):
            return None

        file.seek(0)

        workbook = openpyxl.load_workbook(
            file,
            data_only=True,
            read_only=False
        )

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows():
                for cell in row:
                    cell_value = normalize_text(cell.value)

                    if cell_value == "gamme d'essai":
                        current_row = cell.row
                        current_col = cell.column

                        # 1. Chercher la valeur à droite
                        right_cell = sheet.cell(
                            row=current_row,
                            column=current_col + 1
                        ).value

                        if right_cell not in [None, ""]:
                            return str(right_cell).strip()

                        # 2. Si la cellule à droite est vide, chercher en dessous
                        bottom_cell = sheet.cell(
                            row=current_row + 1,
                            column=current_col
                        ).value

                        if bottom_cell not in [None, ""]:
                            return str(bottom_cell).strip()

                        # 3. Si besoin, chercher diagonale bas-droite
                        bottom_right_cell = sheet.cell(
                            row=current_row + 1,
                            column=current_col + 1
                        ).value

                        if bottom_right_cell not in [None, ""]:
                            return str(bottom_right_cell).strip()

        return None

    except Exception as e:
        print("Erreur extraction nom_gamme:", e)
        return None

    finally:
        try:
            file.seek(0)
        except Exception:
            pass