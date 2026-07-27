import hashlib
import os
import re
import time

from django.core.cache import cache
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string


PARSE_CACHE_TIMEOUT = 6 * 60 * 60
PARSE_LOCK_TIMEOUT = 2 * 60
PARSE_LOCK_WAIT_SECONDS = 30


def clean_html(value):
    """Strip HTML tags and normalize whitespace."""
    if not isinstance(value, str):
        return str(value) if (value is not None and value != "") else ""
    return re.sub(r"<[^>]+>", "", value).strip()



STEP_TYPE_COLOR = {
    "CI": "beige",   
    "AC": "salmon",   
    "RA": "green",    
    "CO": "grey",    
}

EXCEL_BG_COLOR = {
    "FF87CEFA": "ev",
    "FFF5F5DC": "beige",
    "FFF4B084": "salmon",
    "FFA9D08E": "green",
    "FFD9D9D9": "grey",
}

COTATION_OPTIONS = ["A_coter", "OK", "NOK_mineur", "NOK", "Non_coté"]

# Columns to expose to the frontend (display order)
COLUMNS_WANTED = [
    "Nom (Steps)",
    "Désignation EN(EV/STEX)",
    "Désignation EN(STEP)",
    "Description(STEP)",
    "Résultat attendu (Exigence)",
    "Résultat mesuré (Résultats)",
    "Cotation (Résultats)",
    "Commentaire (Résultats)",
    "Image  (Média)",
    "ETAT",
]


def _get_rgb(cell):
    """Return the ARGB hex string of a cell's fill, or None."""
    try:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.type == "rgb":
            return fill.fgColor.rgb
    except Exception:
        pass
    return None


def parse_gamme(file_path):
    wb = None
    try:
        wb = load_workbook(
            file_path,
            read_only=False,
            keep_vba=False,
            keep_links=False,
        )
        ws = wb.active

        # ── 1. Find the header row ─────────────────────────────────────────
        header_row_idx = None
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), 1):
            for cell in row:
                if cell.value and str(cell.value).strip() == "Nom (Steps)":
                    header_row_idx = r_idx
                    break
            if header_row_idx:
                break

        if not header_row_idx:
            return {"error": "Header row with 'Nom (Steps)' not found"}

        # ── 2. Build column index: cleaned_name → 0-based index ───────────
        header_cells = next(ws.iter_rows(
            min_row=header_row_idx, max_row=header_row_idx
        ))
        col_idx = {}
        for cell in header_cells:
            if cell.value:
                name = clean_html(str(cell.value))
                col_idx[name] = cell.column - 1   # 0-based

        # Resolve the columns we want (only those that exist)
        columns = [c for c in COLUMNS_WANTED if c in col_idx]

        # ── 3. Build cotation-eligible rows from data validation ───────────
        #   Primary truth: the Excel data validation on the Cotation column.
        #   This is file-specific and beats everything else.
        cotation_rows = set()
        for dv in ws.data_validations.dataValidation:
            for part in str(dv.sqref).split():
                try:
                    if ":" in part:
                        start, end = part.split(":")
                        _, rs = coordinate_from_string(start)
                        _, re_ = coordinate_from_string(end)
                        for r in range(rs, re_ + 1):
                            cotation_rows.add(r)
                    else:
                        _, r = coordinate_from_string(part)
                        cotation_rows.add(r)
                except Exception:
                    pass

        # ── 4. Parse data rows ─────────────────────────────────────────────
        blocs = []
        current_bloc = None

        for row in ws.iter_rows(min_row=header_row_idx + 1):
            a_cell = row[0]
            step = clean_html(str(a_cell.value)).strip() if a_cell.value else ""
            if not step or step == "TES-X":
                continue

            row_num = a_cell.row
            raw_rgb = _get_rgb(a_cell) or "00000000"

            # ── Determine row color ──────────────────────────────────────
            # Priority 1: Type de step column (most reliable semantic signal)
            type_idx = col_idx.get("Type de step")
            step_type = ""
            if type_idx is not None and row[type_idx].value:
                step_type = str(row[type_idx].value).strip().upper()

            if step_type in STEP_TYPE_COLOR:
                row_color = STEP_TYPE_COLOR[step_type]
            else:
                # Priority 2: actual cell background color
                row_color = EXCEL_BG_COLOR.get(raw_rgb, "beige")

            # ── Cotation eligibility ─────────────────────────────────────
            # Priority 1: data validation presence (exact, file-specific)
            # Priority 2: Type de step == "RA" (semantic, works cross-file)
            # Priority 3: row color == "green" (visual fallback)
            has_cotation = (
                row_num in cotation_rows
                or step_type == "RA"
                or row_color == "green"
            )

            # ── EV row ───────────────────────────────────────────────────
            if step.startswith("EV"):
                if current_bloc:
                    blocs.append(current_bloc)

                ev_cells = []
                for col in columns:
                    idx = col_idx.get(col)
                    val = clean_html(str(row[idx].value)) if (idx is not None and row[idx].value) else ""
                    ev_cells.append({"type": "text", "field": col, "value": val})

                current_bloc = {
                    "title": step,
                    "ev_row": ev_cells,
                    "rows": [],
                }

            # ── STEP row ─────────────────────────────────────────────────
            elif step.startswith("STEP") and current_bloc:
                row_cells = []
                for col in columns:
                    idx = col_idx.get(col)
                    raw_val = row[idx].value if idx is not None else None

                    if col == "Cotation (Résultats)":
                        if has_cotation:
                            saved = clean_html(str(raw_val)).strip() if raw_val else ""
                            row_cells.append({
                                "type": "select",
                                "field": col,
                                "value": saved if saved in COTATION_OPTIONS else "A_coter",
                                "options": COTATION_OPTIONS,
                            })
                        else:
                            row_cells.append({
                                "type": "text",
                                "field": col,
                                "value": "",
                            })
                    else:
                        row_cells.append({
                            "type": "text",
                            "field": col,
                            "value": clean_html(str(raw_val)) if raw_val else "",
                        })

                current_bloc["rows"].append({
                    "cells": row_cells,
                    "color": row_color,   # "beige" | "salmon" | "green" | "grey"
                    "step_type": step_type,
                })

        if current_bloc:
            blocs.append(current_bloc)

        return {"colonnes": columns, "blocs": blocs}

    except Exception as e:
        import traceback
        print("❌ ERREUR:", traceback.format_exc())
        return {"error": str(e)}
    finally:
        if wb is not None:
            wb.close()


def _parse_cache_key(file_path):
    stats = os.stat(file_path)
    fingerprint = (
        f"v2:{os.path.abspath(file_path)}:{stats.st_size}:"
        f"{stats.st_mtime_ns}"
    )
    digest = hashlib.sha256(fingerprint.encode("utf-8")).hexdigest()
    return f"gamme-parse:{digest}"


def parse_gamme_cached(file_path):
    try:
        cache_key = _parse_cache_key(file_path)
        cached = cache.get(cache_key)
    except Exception:
        return parse_gamme(file_path)

    if cached is not None:
        return cached

    lock_key = f"{cache_key}:lock"
    try:
        owns_lock = cache.add(lock_key, "1", PARSE_LOCK_TIMEOUT)
    except Exception:
        return parse_gamme(file_path)

    if owns_lock:
        try:
            parsed = parse_gamme(file_path)
            if not parsed.get("error"):
                cache.set(cache_key, parsed, PARSE_CACHE_TIMEOUT)
            return parsed
        finally:
            try:
                cache.delete(lock_key)
            except Exception:
                pass

    deadline = time.monotonic() + PARSE_LOCK_WAIT_SECONDS
    while time.monotonic() < deadline:
        time.sleep(0.1)
        try:
            cached = cache.get(cache_key)
        except Exception:
            break
        if cached is not None:
            return cached

    return parse_gamme(file_path)
