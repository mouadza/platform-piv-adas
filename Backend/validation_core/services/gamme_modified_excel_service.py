import os
import re
import unicodedata
from collections import defaultdict
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from validation_core.models.measured_result_comment import StepMeasuredResultComment
from validation_core.models.results import StepValidation
from validation_core.services.gamme_validation_dates import are_all_gamme_evs_validated


def clean_text(value):
    if value is None:
        return ""

    return re.sub(r"<[^>]+>", "", str(value)).strip()


def normalize_text(value):
    text = clean_text(value).lower()

    replacements = {
        "\u00c3\u00a9": "e",  # UTF-8 "é" previously decoded as Latin-1
        "\u00c3\u00a8": "e",
        "\u00c3\u00aa": "e",
        "\u00c3\u2030": "e",
        "\u00c3\u00a0": "a",
        "\u00c3\u00a2": "a",
        "\u00c3\u00a7": "c",
        "\u00c3\u00b4": "o",
        "\u00c3\u00bb": "u",
    }

    for source, target in replacements.items():
        text = text.replace(source.lower(), target)

    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9]+", " ", text)

    return " ".join(text.split())


def is_step_header(value):
    normalized = normalize_text(value)

    return "nom" in normalized and "steps" in normalized


def is_measured_result_header(value):
    normalized = normalize_text(value)

    return "resultat" in normalized and "mesur" in normalized


def is_cotation_header(value):
    normalized = normalize_text(value)

    return "cotation" in normalized


def is_comment_header(value):
    normalized = normalize_text(value)

    return "commentaire" in normalized and "result" in normalized


def find_header_row(worksheet):
    for row in worksheet.iter_rows(min_row=1, max_row=min(30, worksheet.max_row)):
        for cell in row:
            if is_step_header(cell.value):
                return cell.row

    raise ValueError("Impossible de trouver la ligne d'en-tete du fichier gamme.")


def find_column(header_cells, predicate):
    for cell in header_cells:
        if predicate(cell.value):
            return cell.column

    return None


def build_latest_validations(gamme):
    latest = {}

    validations = StepValidation.objects.filter(gamme=gamme).order_by(
        "-created_at",
        "-id",
    )

    for validation in validations:
        key = (validation.ev_code or "", validation.step_code)

        if key not in latest:
            latest[key] = validation

    return latest


def build_measured_comments(gamme):
    comments_by_step = defaultdict(list)

    comments = StepMeasuredResultComment.objects.filter(gamme=gamme).order_by(
        "created_at",
        "id",
    )

    for comment in comments:
        key = (comment.ev_code or "", comment.step_code)
        comments_by_step[key].append(comment.commentaire)

    return comments_by_step


def normalize_cotation_for_excel(cotation):
    if cotation == "NOK_mineur":
        return "NOK_mineur"

    return cotation or ""


def write_wrapped_cell(worksheet, row, column, value):
    cell = worksheet.cell(row=row, column=column)
    cell.value = value
    cell.alignment = Alignment(
        horizontal=cell.alignment.horizontal,
        vertical=cell.alignment.vertical,
        text_rotation=cell.alignment.text_rotation,
        wrap_text=True,
        shrink_to_fit=cell.alignment.shrink_to_fit,
        indent=cell.alignment.indent,
    )


def generate_modified_gamme_excel(gamme):
    if not gamme.fichier_gamme:
        raise ValueError("Cette gamme n'a pas de fichier Excel associe.")

    if not are_all_gamme_evs_validated(gamme):
        raise ValueError(
            "Le fichier Excel modifie est disponible seulement lorsque tous les EV sont valides."
        )

    source_path = gamme.fichier_gamme.path
    extension = os.path.splitext(source_path)[1].lower()

    if extension not in [".xlsx", ".xlsm"]:
        raise ValueError("Seuls les fichiers .xlsx et .xlsm peuvent etre exportes.")

    workbook = load_workbook(
        source_path,
        keep_vba=extension == ".xlsm",
        keep_links=False,
    )
    worksheet = workbook.active

    header_row = find_header_row(worksheet)
    header_cells = next(
        worksheet.iter_rows(min_row=header_row, max_row=header_row)
    )

    measured_result_col = find_column(header_cells, is_measured_result_header)
    cotation_col = find_column(header_cells, is_cotation_header)
    comment_col = find_column(header_cells, is_comment_header)

    if not any([measured_result_col, cotation_col, comment_col]):
        raise ValueError(
            "Impossible de trouver les colonnes Resultat mesure, Cotation ou Commentaire."
        )

    latest_validations = build_latest_validations(gamme)
    measured_comments = build_measured_comments(gamme)

    current_ev_code = None

    for row in worksheet.iter_rows(min_row=header_row + 1):
        first_cell = row[0]
        row_label = clean_text(first_cell.value)

        if not row_label:
            continue

        if row_label.startswith("EV"):
            current_ev_code = row_label
            continue

        if not row_label.startswith("STEP") or not current_ev_code:
            continue

        key = (current_ev_code, row_label)
        validation = latest_validations.get(key)
        row_number = first_cell.row

        if validation and cotation_col:
            worksheet.cell(
                row=row_number,
                column=cotation_col,
                value=normalize_cotation_for_excel(validation.cotation),
            )

        if validation and comment_col:
            write_wrapped_cell(
                worksheet,
                row_number,
                comment_col,
                validation.commentaire or "",
            )

        if measured_result_col and key in measured_comments:
            write_wrapped_cell(
                worksheet,
                row_number,
                measured_result_col,
                "\n".join(filter(None, measured_comments[key])),
            )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    workbook.close()

    original_name = (
        gamme.original_filename
        or os.path.basename(gamme.fichier_gamme.name)
        or f"gamme_{gamme.id}{extension}"
    )
    base_name, original_extension = os.path.splitext(original_name)
    output_extension = original_extension or extension

    return {
        "content": output,
        "filename": f"{base_name}_modifie{output_extension}",
        "extension": output_extension.lower(),
    }
