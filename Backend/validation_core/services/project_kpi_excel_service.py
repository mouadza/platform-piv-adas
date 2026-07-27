from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from validation_core.services.gamme_parse_storage import get_or_parse_gamme_data


KPI_LEFT_COLUMNS_COUNT = 10
EV_RESULT_START_COLUMN = 2
EV_RESULT_COLUMNS_COUNT = 4
GAMME_KPI_START_COLUMN = 6
TIMELINE_START_COLUMN = KPI_LEFT_COLUMNS_COUNT + 1

COLORS = {
    "dark_blue": "1F4E79",
    "header_blue": "244061",
    "grid": "D9E2F3",
    "white": "FFFFFF",
    "light_grey": "F8FAFC",
    "text": "111827",
    "muted_text": "475569",
    "ok": "00B050",
    "nok_minor": "F79646",
    "nok": "FF0000",
    "non_cote": "BFBFBF",
    "a_coter": "000000",
}

STATUS_ROWS = [
    ("OK", "okSteps", "okPercent", COLORS["ok"], COLORS["text"]),
    (
        "NOK Mineur",
        "minorSteps",
        "minorPercent",
        COLORS["nok_minor"],
        COLORS["text"],
    ),
    ("NOK", "nokSteps", "nokPercent", COLORS["nok"], COLORS["text"]),
    (
        "Non cote",
        "nonCoteSteps",
        "nonCotePercent",
        COLORS["non_cote"],
        COLORS["text"],
    ),
    (
        "A coter",
        "aCoterSteps",
        "aCoterPercent",
        COLORS["a_coter"],
        COLORS["white"],
    ),
]

EV_RESULT_ROWS = [
    ("OK", "OK", COLORS["ok"], COLORS["text"]),
    (
        "NOK Mineur",
        "NOK_mineur",
        COLORS["nok_minor"],
        COLORS["text"],
    ),
    ("NOK", "NOK", COLORS["nok"], COLORS["text"]),
    ("En cours", "IN_PROGRESS", COLORS["grid"], COLORS["text"]),
]


def _fill(color):
    return PatternFill(fill_type="solid", fgColor=color)


def _border(color=None):
    side = Side(style="thin", color=color or COLORS["grid"])
    return Border(top=side, left=side, bottom=side, right=side)


def _percent(value, total):
    if not total:
        return 0
    return round((float(value or 0) / float(total)) * 100, 1)


def _normalize_cotation(value):
    normalized = str(value or "").strip()
    if not normalized or normalized == "A_coter":
        return "A_coter"
    if normalized.startswith("Non_cot"):
        return "Non_cote"
    if normalized.lower() == "nok mineur":
        return "NOK_mineur"
    return normalized


def _summarize_cotations(cotations):
    normalized = [_normalize_cotation(value) for value in cotations]
    total = len(normalized)
    ok = normalized.count("OK")
    nok = normalized.count("NOK")
    minor = normalized.count("NOK_mineur")
    non_cote = normalized.count("Non_cote")
    a_coter = normalized.count("A_coter")
    validated = total - a_coter

    return {
        "total": total,
        "validated": validated,
        "ok": ok,
        "nok": nok,
        "minor": minor,
        "nonCote": non_cote,
        "aCoter": a_coter,
        "completionPercent": _percent(validated, total),
        "okPercent": _percent(ok, total),
        "nokPercent": _percent(nok, total),
        "minorPercent": _percent(minor, total),
        "nonCotePercent": _percent(non_cote, total),
        "aCoterPercent": _percent(a_coter, total),
    }


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).date()
    except (TypeError, ValueError):
        return None


def _gamme_name(gamme):
    return (
        gamme.nom_gamme
        or gamme.original_filename
        or gamme.nom
        or f"Gamme {gamme.id}"
    )


def _ev_code(bloc):
    for cell in bloc.get("ev_row") or []:
        if cell.get("value"):
            return str(cell["value"])
    return "UNKNOWN_EV"


def _step_code(row):
    for cell in row.get("cells") or []:
        if cell.get("field") == "Nom (Steps)":
            return str(cell.get("value") or "-")
    return "-"


def _has_cotation(row):
    return any(
        cell.get("type") == "select"
        and str(cell.get("field") or "").startswith("Cotation (")
        for cell in row.get("cells") or []
    )


def _build_date_range(gamme, validations):
    validation_dates = sorted(
        value
        for value in (_as_date(item.created_at) for item in validations)
        if value
    )
    if not validation_dates:
        return None, None, 0

    duration = max(int(gamme.nombre_jours or 1), 1)
    start = _as_date(gamme.date_debut) or validation_dates[0]
    fallback_end = start + timedelta(days=duration - 1)
    end = _as_date(gamme.date_fin) or validation_dates[-1] or fallback_end
    duration_days = max((end - start).days + 1, 1)
    return start, end, duration_days


def build_project_kpi_data(project, progress_callback=None):
    from validation_core.models import StepValidation

    gammes = list(
        project.gammes.select_related("vehicule").order_by("ordre", "id")
    )
    data = []
    total_gammes = len(gammes)

    for index, gamme in enumerate(gammes, start=1):
        if not gamme.fichier_gamme:
            parsed = {"blocs": []}
        else:
            parsed = get_or_parse_gamme_data(gamme)
            if parsed.get("error"):
                raise ValueError(
                    f"Impossible de parser la gamme {_gamme_name(gamme)}: "
                    f"{parsed['error']}"
                )

        validations = list(
            StepValidation.objects.filter(gamme=gamme).order_by(
                "created_at", "id"
            )
        )
        latest = {}
        for validation in validations:
            latest[(validation.ev_code or "", validation.step_code)] = validation
        latest_validations = list(latest.values())

        ev_map = {}
        all_cotations = []

        for bloc in parsed.get("blocs") or []:
            ev_code = _ev_code(bloc)
            ev_map.setdefault(ev_code, [])

            for row in bloc.get("rows") or []:
                if not _has_cotation(row):
                    continue
                step_code = _step_code(row)
                validation = latest.get((ev_code, step_code))
                cotation = _normalize_cotation(
                    getattr(validation, "cotation", None)
                )
                ev_map[ev_code].append(cotation)
                all_cotations.append(cotation)

        summary = _summarize_cotations(all_cotations)
        start, end, duration_days = _build_date_range(
            gamme,
            latest_validations,
        )
        ev_stats = []

        for ev_code, cotations in ev_map.items():
            ev_stats.append(
                {
                    "evCode": ev_code,
                    **_summarize_cotations(cotations),
                }
            )

        data.append(
            {
                "gammeId": gamme.id,
                "name": _gamme_name(gamme),
                "progress": summary["completionPercent"],
                "totalSteps": summary["total"],
                "validatedSteps": summary["validated"],
                "okSteps": summary["ok"],
                "nokSteps": summary["nok"],
                "minorSteps": summary["minor"],
                "nonCoteSteps": summary["nonCote"],
                "aCoterSteps": summary["aCoter"],
                "completionPercent": summary["completionPercent"],
                "okPercent": summary["okPercent"],
                "nokPercent": summary["nokPercent"],
                "minorPercent": summary["minorPercent"],
                "nonCotePercent": summary["nonCotePercent"],
                "aCoterPercent": summary["aCoterPercent"],
                "evStats": ev_stats,
                "startDate": start,
                "endDate": end,
                "durationDays": duration_days,
            }
        )

        if progress_callback:
            progress_callback(index, total_gammes)

    return data


def _with_percentages(summary):
    total = int(summary.get("totalSteps") or 0)
    ok = int(summary.get("okSteps") or 0)
    nok = int(summary.get("nokSteps") or 0)
    minor = int(summary.get("minorSteps") or 0)
    non_cote = int(summary.get("nonCoteSteps") or 0)
    a_coter = int(summary.get("aCoterSteps") or 0)
    validated = total - a_coter
    return {
        **summary,
        "totalSteps": total,
        "validatedSteps": validated,
        "okSteps": ok,
        "nokSteps": nok,
        "minorSteps": minor,
        "nonCoteSteps": non_cote,
        "aCoterSteps": a_coter,
        "completionPercent": _percent(validated, total),
        "okPercent": _percent(ok, total),
        "nokPercent": _percent(nok, total),
        "minorPercent": _percent(minor, total),
        "nonCotePercent": _percent(non_cote, total),
        "aCoterPercent": _percent(a_coter, total),
    }


def _summarize_project(kpi_data):
    summary = {
        "totalSteps": 0,
        "okSteps": 0,
        "nokSteps": 0,
        "minorSteps": 0,
        "nonCoteSteps": 0,
        "aCoterSteps": 0,
    }
    for gamme in kpi_data:
        for target, source in (
            ("totalSteps", "totalSteps"),
            ("okSteps", "okSteps"),
            ("nokSteps", "nokSteps"),
            ("minorSteps", "minorSteps"),
            ("nonCoteSteps", "nonCoteSteps"),
            ("aCoterSteps", "aCoterSteps"),
        ):
            summary[target] += int(gamme.get(source) or 0)
    return _with_percentages(summary)


def _normalize_ev_summary(ev):
    return _with_percentages(
        {
            "totalSteps": ev.get("total"),
            "okSteps": ev.get("ok"),
            "nokSteps": ev.get("nok"),
            "minorSteps": ev.get("minor"),
            "nonCoteSteps": ev.get("nonCote"),
            "aCoterSteps": ev.get("aCoter"),
        }
    )


def _ev_global_result(summary):
    if not summary["totalSteps"] or summary["aCoterSteps"]:
        return "IN_PROGRESS"
    if summary["nokSteps"]:
        return "NOK"
    if summary["minorSteps"]:
        return "NOK_mineur"
    return "OK"


def _ev_occurrences(kpi_data):
    occurrences = []
    for gamme in kpi_data:
        for ev in gamme.get("evStats") or []:
            summary = _normalize_ev_summary(ev)
            occurrences.append(
                {
                    "gammeName": gamme["name"],
                    "evCode": ev["evCode"],
                    "result": _ev_global_result(summary),
                    **summary,
                }
            )
    return occurrences


def _ev_result_rows(occurrences):
    total = len(occurrences)
    counts = dict.fromkeys((result for _, result, _, _ in EV_RESULT_ROWS), 0)
    for occurrence in occurrences:
        counts[occurrence["result"]] = counts.get(occurrence["result"], 0) + 1
    return [
        {
            "label": label,
            "count": counts[result],
            "percent": _percent(counts[result], total),
            "fill": fill,
            "font_color": font_color,
        }
        for label, result, fill, font_color in EV_RESULT_ROWS
    ]


def _start_of_week(value):
    return value - timedelta(days=value.weekday())


def _end_of_week(value):
    return _start_of_week(value) + timedelta(days=6)


def _timeline_bounds(kpi_data):
    starts = [item["startDate"] for item in kpi_data if item.get("startDate")]
    ends = [item["endDate"] for item in kpi_data if item.get("endDate")]
    if not starts or not ends:
        return None, None, False
    return _start_of_week(min(starts)), _end_of_week(max(ends)), True


def _timeline_units(start, end):
    units = []
    current = start
    while start and end and current <= end:
        units.append((current, min(current + timedelta(days=6), end)))
        current += timedelta(days=7)
    return units


def _style_cell(
    cell,
    *,
    fill=None,
    font_color=None,
    bold=False,
    size=None,
    horizontal="center",
    border_color=None,
):
    if fill:
        cell.fill = _fill(fill)
    cell.border = _border(border_color)
    cell.font = Font(
        bold=bold,
        size=size,
        color=font_color or COLORS["text"],
    )
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical="center",
        wrap_text=True,
    )


def _style_percent(cell, value, number_format="0.00%"):
    cell.value = float(value or 0) / 100
    cell.number_format = number_format
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _merge(ws, start_row, start_col, end_row, end_col):
    if start_row != end_row or start_col != end_col:
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=end_row,
            end_column=end_col,
        )


def _title_row(ws, row_number, total_columns, value):
    for column in range(1, total_columns + 1):
        ws.cell(row_number, column).fill = _fill(COLORS["dark_blue"])
    _merge(ws, row_number, 1, row_number, total_columns)
    cell = ws.cell(row_number, 1, value)
    cell.font = Font(bold=True, size=16, color=COLORS["white"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_number].height = 28


def _info_row(ws, row_number, total_columns, value):
    for column in range(1, total_columns + 1):
        cell = ws.cell(row_number, column)
        cell.fill = _fill(COLORS["light_grey"])
        cell.border = _border()
    _merge(ws, row_number, 1, row_number, total_columns)
    cell = ws.cell(row_number, 1, value)
    cell.font = Font(bold=True, color=COLORS["muted_text"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_number].height = 22


def _summary_band(ws, row, start_col, end_col, title):
    for column in range(start_col, end_col + 1):
        cell = ws.cell(row, column)
        cell.fill = _fill(COLORS["dark_blue"])
        cell.border = _border(COLORS["white"])
    _merge(ws, row, start_col, row, end_col)
    cell = ws.cell(row, start_col, title)
    cell.font = Font(bold=True, color=COLORS["white"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24


def _summary_header(ws, row, start_col, end_col, labels):
    for column in range(start_col, end_col + 1):
        cell = ws.cell(row, column, labels[column - start_col] if column - start_col < len(labels) else "")
        _style_cell(
            cell,
            fill=COLORS["header_blue"],
            font_color=COLORS["white"],
            bold=True,
            border_color=COLORS["white"],
        )
    ws.row_dimensions[row].height = 24


def _summary_rows(ws, start_row, start_col, end_col, rows, total):
    for index, item in enumerate(rows):
        row_number = start_row + index
        for column in range(start_col, end_col + 1):
            _style_cell(ws.cell(row_number, column))
        label = ws.cell(row_number, start_col, item["label"])
        label.fill = _fill(item["fill"])
        label.font = Font(bold=True, color=item["font_color"])
        ws.cell(row_number, start_col + 1, item["count"])
        _style_percent(ws.cell(row_number, start_col + 2), item["percent"])

    if rows:
        total_column = start_col + 3
        end_row = start_row + len(rows) - 1
        _merge(ws, start_row, total_column, end_row, total_column)
        total_cell = ws.cell(start_row, total_column, total)
        _style_cell(total_cell, bold=True, size=11)


def _global_cotation_rows(kpi_data):
    summary = _summarize_project(kpi_data)
    return [
        {
            "label": label,
            "count": summary[count_field],
            "percent": summary[percent_field],
            "fill": fill,
            "font_color": font_color,
        }
        for label, count_field, percent_field, fill, font_color in STATUS_ROWS
    ]


def _top_summaries(ws, total_columns, project_name, kpi_data, start_row):
    left_start, left_end = 1, min(8, total_columns)
    right_start, right_end = min(10, total_columns), total_columns
    ev_occurrences = _ev_occurrences(kpi_data)
    ev_rows = _ev_result_rows(ev_occurrences)
    cotation_rows = _global_cotation_rows(kpi_data)
    project_summary = _summarize_project(kpi_data)

    _summary_band(
        ws,
        start_row,
        left_start,
        left_end,
        f"Resultat EV global du projet - {project_name}",
    )
    _summary_header(
        ws,
        start_row + 1,
        left_start,
        left_end,
        ["Resultat EV", "Nombre", "Pourcentage", "Total"],
    )
    _summary_rows(
        ws,
        start_row + 2,
        left_start,
        left_end,
        ev_rows,
        len(ev_occurrences),
    )

    _summary_band(
        ws,
        start_row,
        right_start,
        right_end,
        f"Cotations globales du projet - {project_name}",
    )
    _summary_header(
        ws,
        start_row + 1,
        right_start,
        right_end,
        ["Cotation", "Nombre", "Pourcentage", "Total"],
    )
    _summary_rows(
        ws,
        start_row + 2,
        right_start,
        right_end,
        cotation_rows,
        project_summary["totalSteps"],
    )
    return start_row + max(len(ev_rows), len(cotation_rows)) + 1


def _section_headers(ws, row_number, total_columns, units, has_bounds):
    for column in range(1, total_columns + 1):
        _style_cell(
            ws.cell(row_number, column),
            fill=COLORS["dark_blue"],
            font_color=COLORS["white"],
            bold=True,
            border_color=COLORS["white"],
        )
    ws.cell(row_number, 1, "KPI par EV")
    ws.cell(row_number, GAMME_KPI_START_COLUMN, "KPI par gamme")
    _merge(
        ws,
        row_number,
        1,
        row_number,
        EV_RESULT_START_COLUMN + EV_RESULT_COLUMNS_COUNT - 1,
    )
    _merge(
        ws,
        row_number,
        GAMME_KPI_START_COLUMN,
        row_number,
        KPI_LEFT_COLUMNS_COUNT,
    )
    if has_bounds:
        ws.cell(row_number, TIMELINE_START_COLUMN, "Calendrier par semaine")
        _merge(
            ws,
            row_number,
            TIMELINE_START_COLUMN,
            row_number,
            total_columns,
        )
    ws.row_dimensions[row_number].height = 22

    header_row = row_number + 1
    labels = [
        "Gamme",
        "Resultat EV",
        "%",
        "Nb",
        "Total",
        "Cotation",
        "%",
        "Nb",
        "Total",
        "% Avancement",
    ]
    for column in range(1, total_columns + 1):
        if column <= len(labels):
            value = labels[column - 1]
        else:
            unit_index = column - TIMELINE_START_COLUMN
            value = units[unit_index][0] if unit_index < len(units) else ""
        cell = ws.cell(header_row, column, value)
        _style_cell(
            cell,
            fill=COLORS["header_blue"],
            font_color=COLORS["white"],
            bold=True,
            border_color=COLORS["white"],
        )
        if column >= TIMELINE_START_COLUMN:
            cell.number_format = "dd/mm/yyyy"
    ws.row_dimensions[header_row].height = 24
    return header_row


def _display_progress(gamme):
    progress = (
        float(gamme.get("okPercent") or 0)
        + float(gamme.get("nokPercent") or 0)
        + float(gamme.get("minorPercent") or 0)
    )
    return max(0, min(progress, 100))


def _add_gamme_block(ws, start_row, gamme, units, total_columns):
    ev_occurrences = []
    for ev in gamme.get("evStats") or []:
        summary = _normalize_ev_summary(ev)
        ev_occurrences.append(
            {
                "result": _ev_global_result(summary),
                **summary,
            }
        )
    ev_rows = _ev_result_rows(ev_occurrences)

    for index, status in enumerate(STATUS_ROWS):
        row_number = start_row + index
        for column in range(1, total_columns + 1):
            _style_cell(ws.cell(row_number, column), horizontal="left")
        ws.row_dimensions[row_number].height = 22

        if index < len(ev_rows):
            ev = ev_rows[index]
            cell = ws.cell(row_number, EV_RESULT_START_COLUMN, ev["label"])
            cell.fill = _fill(ev["fill"])
            cell.font = Font(bold=True, color=ev["font_color"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            _style_percent(
                ws.cell(row_number, EV_RESULT_START_COLUMN + 1),
                ev["percent"],
            )
            count = ws.cell(
                row_number,
                EV_RESULT_START_COLUMN + 2,
                ev["count"],
            )
            count.alignment = Alignment(horizontal="center", vertical="center")

        label, count_field, percent_field, fill, font_color = status
        status_cell = ws.cell(row_number, GAMME_KPI_START_COLUMN, label)
        status_cell.fill = _fill(fill)
        status_cell.font = Font(bold=True, color=font_color)
        status_cell.alignment = Alignment(horizontal="left", vertical="center")
        _style_percent(
            ws.cell(row_number, GAMME_KPI_START_COLUMN + 1),
            gamme.get(percent_field),
        )
        count_cell = ws.cell(
            row_number,
            GAMME_KPI_START_COLUMN + 2,
            int(gamme.get(count_field) or 0),
        )
        count_cell.alignment = Alignment(horizontal="center", vertical="center")

    end_row = start_row + len(STATUS_ROWS) - 1
    _merge(ws, start_row, 1, end_row, 1)
    _merge(
        ws,
        start_row,
        EV_RESULT_START_COLUMN + 3,
        start_row + max(len(ev_rows), 1) - 1,
        EV_RESULT_START_COLUMN + 3,
    )
    _merge(
        ws,
        start_row,
        GAMME_KPI_START_COLUMN + 3,
        end_row,
        GAMME_KPI_START_COLUMN + 3,
    )
    _merge(
        ws,
        start_row,
        GAMME_KPI_START_COLUMN + 4,
        end_row,
        GAMME_KPI_START_COLUMN + 4,
    )

    gamme_cell = ws.cell(
        start_row,
        1,
        f"{gamme['name']}\n({int(gamme.get('totalSteps') or 0)})",
    )
    _style_cell(gamme_cell, bold=True)

    ev_total = ws.cell(
        start_row,
        EV_RESULT_START_COLUMN + 3,
        len(ev_occurrences),
    )
    _style_cell(ev_total, bold=True, size=11)

    total_cell = ws.cell(
        start_row,
        GAMME_KPI_START_COLUMN + 3,
        int(gamme.get("totalSteps") or 0),
    )
    _style_cell(total_cell, bold=True, size=11)

    progress_cell = ws.cell(start_row, GAMME_KPI_START_COLUMN + 4)
    _style_cell(progress_cell, bold=True, size=11)
    _style_percent(progress_cell, _display_progress(gamme), "0.0%")
    progress_cell.font = Font(bold=True, size=11, color=COLORS["text"])

    start = gamme.get("startDate")
    end = gamme.get("endDate")
    if start and end:
        for index, (unit_start, unit_end) in enumerate(units):
            if start <= unit_end and end >= unit_start:
                cell = ws.cell(start_row, TIMELINE_START_COLUMN + index, "")
                cell.fill = _fill(COLORS["dark_blue"])
                cell.border = _border(COLORS["white"])
    return end_row


def create_project_kpi_workbook(project_name, kpi_data):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "KPI Projet"

    project_start, project_end, has_bounds = _timeline_bounds(kpi_data)
    units = _timeline_units(project_start, project_end)
    timeline_columns = TIMELINE_START_COLUMN + len(units) - 1
    total_columns = max(timeline_columns, 17)

    _title_row(ws, 1, total_columns, f"KPI Projet - {project_name}")
    planning = (
        f"Planning par gamme : {project_start.strftime('%d/%m/%Y')} - "
        f"{project_end.strftime('%d/%m/%Y')}"
        if has_bounds
        else "Planning par gamme : aucune gamme commencee"
    )
    _info_row(ws, 2, total_columns, planning)
    ws.row_dimensions[3].height = 10

    summary_end = _top_summaries(ws, total_columns, project_name, kpi_data, 4)
    spacer_row = summary_end + 1
    ws.row_dimensions[spacer_row].height = 10
    section_row = spacer_row + 1
    header_row = _section_headers(
        ws,
        section_row,
        total_columns,
        units,
        has_bounds,
    )

    current_row = header_row + 1
    for gamme in kpi_data:
        current_row = _add_gamme_block(
            ws,
            current_row,
            gamme,
            units,
            total_columns,
        ) + 1

    widths = [22, 18, 12, 14, 10, 18, 10, 10, 10, 14, 12, 12, 12, 12, 12, 12, 10]
    for column in range(1, total_columns + 1):
        ws.column_dimensions[get_column_letter(column)].width = (
            widths[column - 1] if column <= len(widths) else 12
        )

    ws.freeze_panes = ws.cell(header_row + 1, TIMELINE_START_COLUMN)
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(total_columns)}"
        f"{max(header_row, current_row - 1)}"
    )
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_format.defaultRowHeight = 22

    workbook.properties.creator = "RepProject"
    workbook.properties.title = f"KPI Projet - {project_name}"
    workbook.properties.subject = project_name
    workbook.properties.created = datetime.now()
    return workbook


def generate_project_kpi_excel(project, progress_callback=None):
    kpi_data = build_project_kpi_data(project, progress_callback)
    workbook = create_project_kpi_workbook(project.nom_projet, kpi_data)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
